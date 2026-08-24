import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import re

wb = openpyxl.Workbook()

# ── Styles ──
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="4472C4")
epic_font = Font(bold=True, size=12, color="1F4E79")
epic_fill = PatternFill("solid", fgColor="D6E4F0")
wrap = Alignment(wrap_text=True, vertical="top")
thin = Side(style="thin")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def write_table(ws, headers, rows, col_widths):
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        c.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = wrap
            c.border = border

def read_md(path):
    with open(path, "r", encoding="utf-8") as f:
        return f.read()

def parse_epics(text):
    epics = []
    current_epic = None
    header_line = None
    for line in text.splitlines():
        m = re.match(r"^## (EP - \d+ .+)$", line)
        if m:
            current_epic = m.group(1)
            header_line = None
            continue
        if current_epic and re.match(r"^\| ID ", line):
            header_line = line
            continue
        if current_epic and header_line and re.match(r"^\| DC", line):
            cells = [c.strip() for c in line.split("|")[1:-1]]
            if len(cells) >= 6:
                epics.append((current_epic, cells[0], cells[2], cells[3], cells[4], cells[5]))
    return epics

text = read_md(
    r"C:\Users\thanh\Desktop\ERP_Cloudaz\docs\thu-hoi-cong-no\srs\sub2-debt-collection\ProductBacklog_Debt_Collection_2026-08-20.md"
)
data = parse_epics(text)

# ── Sheet 1: Full Product Backlog ──
ws1 = wb.active
ws1.title = "Product Backlog"
headers = ["Epic", "ID", "Feature", "User Story (Short)", "User Story (Detail)", "Acceptance Criteria"]
widths = [32, 10, 28, 40, 55, 55]
rows = [(e[0], e[1], e[2], e[3], e[4], e[5]) for e in data]
write_table(ws1, headers, rows, widths)

# ── Sheet 2: Theo Epic ──
ws2 = wb.create_sheet("Theo Epic")
headers2 = ["ID", "Feature", "User Story (Short)", "User Story (Detail)", "Acceptance Criteria"]
widths2 = [10, 28, 40, 55, 55]

current_epic = None
row_idx = 1
for e in data:
    if e[0] != current_epic:
        current_epic = e[0]
        c = ws2.cell(row=row_idx, column=1, value=current_epic)
        c.font = epic_font
        c.fill = epic_fill
        ws2.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=5)
        for ci in range(1, 6):
            ws2.cell(row=row_idx, column=ci).border = border
        row_idx += 1
    for ci, (h, w) in enumerate(zip(headers2, widths2), 1):
        c = ws2.cell(row=row_idx, column=ci, value=e[ci])
        c.alignment = wrap
        c.border = border
        if ci == 1:
            ws2.column_dimensions[get_column_letter(ci)].width = max(
                ws2.column_dimensions[get_column_letter(ci)].width or 0, w
            )
    row_idx += 1

for ci, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(ci)].width = w

# ── Sheet 3: Epic Summary ──
ws3 = wb.create_sheet("Epic Summary")
summary_headers = ["Epic", "Số PBIs", "P0", "P1", "P2"]
summary_widths = [32, 10, 8, 8, 8]

# Count by epic
from collections import Counter, OrderedDict

epic_counts = OrderedDict()
epic_p0 = Counter()
epic_p1 = Counter()

for e in data:
    epic = e[0]
    epic_counts.setdefault(epic, 0)
    epic_counts[epic] += 1

# We don't have priority in the table data, so use the summary table in MD
# Parse the summary table from md
summary_rows = []
in_summary = False
for line in text.splitlines():
    if "| Epic | Số PBIs |" in line:
        in_summary = True
        continue
    if in_summary and re.match(r"^\|[-| ]+\|$", line):
        continue
    if in_summary and re.match(r"^\|", line):
        cells = [c.strip() for c in line.split("|")[1:-1]]
        if len(cells) >= 5 and cells[0] != "Tổng":
            summary_rows.append(cells)
        elif len(cells) >= 5 and cells[0] == "Tổng":
            summary_rows.append(cells)

write_table(ws3, summary_headers, summary_rows, summary_widths)

# ── Sheet 4: Priority Summary ──
ws4 = wb.create_sheet("Priority Summary")
pri_headers = ["Ưu tiên", "Số lượng", "Mô tả"]
pri_widths = [10, 10, 40]

pri_rows = []
in_pri = False
for line in text.splitlines():
    if "| **P0**" in line:
        in_pri = True
    if in_pri and re.match(r"^\|", line):
        cells = [c.strip() for c in line.split("|")[1:-1]]
        if len(cells) >= 3:
            pri_rows.append(cells)

write_table(ws4, pri_headers, pri_rows, pri_widths)

out_path = r"C:\Users\thanh\Desktop\ERP_Cloudaz\docs\thu-hoi-cong-no\srs\sub2-debt-collection\ProductBacklog_Debt_Collection_2026-08-20.xlsx"
wb.save(out_path)
print("OK -> " + out_path)
